#!/usr/bin/env python3
"""export_excel.py — turn ranking results into a reviewable spreadsheet.

    python export_excel.py Stored_results/                    -> rankings.xlsx
    python export_excel.py Stored_results/ --out shortlist.xlsx
    python export_excel.py a_ranking.json b_ranking.json --out both.xlsx

Reads the `*_ranking.json` files `main.py` already wrote, so it makes no API
calls and can be re-run freely.

One sheet per job description, plus a Summary sheet. Screening two hundred
candidates through a web page means two hundred rows nobody scrolls; the same
data in Excel can be sorted, filtered, commented on, and mailed to a hiring
manager who will never install Python.

Every column the pipeline produces travels with the row — matched skills, gaps,
the per-bucket breakdown, the per-skill depth judgements, and the full
explanation — so a reviewer questioning a score can answer it in the same view
rather than opening a JSON file.
"""

from __future__ import annotations

import argparse
import json
import ntpath
import sys
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

FONT = "Arial"

# (header, json key or callable, width, wrap)
COLUMNS: List[Tuple[str, str, int, bool]] = [
    ("Resume",            "resume",      30, False),
    ("JD",                "jd",          26, False),
    ("Score",             "score",        8, False),
    ("Rank",              "rank",         7, False),
    ("Matched Skills",    "matched",     42, True),
    ("Missing Skills",    "missing",     42, True),
    ("Mandatory Skills",  "mandatory",   34, True),
    ("Preferred Skills",  "preferred",   34, True),
    ("Assessment",        "assessment",  52, True),
    ("Explanation",       "explanation", 60, True),
    ("Summary",           "summary",     52, True),
    ("Experience",        "experience",  18, True),
    ("Flags",             "flags",       40, True),
    ("Extraction Conf.",  "conf",        11, False),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT, size=10)
DUP_FONT = Font(name=FONT, size=10, italic=True, color="808080")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# reading


def stem(path: str) -> str:
    """Basename without extension, tolerating Windows separators in the JSON."""
    return ntpath.splitext(ntpath.basename(str(path).replace("/", "\\")))[0]


def basename(path: str) -> str:
    return ntpath.basename(str(path).replace("/", "\\")) or str(path)


def collect(targets: Sequence[str]) -> List[Path]:
    """Every ranking file named, or found under a folder that was named."""
    out: List[Path] = []
    for t in targets:
        p = Path(t)
        if p.is_dir():
            out.extend(sorted(p.rglob("*_ranking.json")))
        elif p.is_file():
            out.append(p)
        else:
            raise SystemExit(f"error: no such file or folder: {t}")
    if not out:
        raise SystemExit("error: no *_ranking.json files found")
    # a folder scan can surface the same file twice
    seen, unique = set(), []
    for p in out:
        key = p.resolve()
        if key not in seen:
            seen.add(key)
            unique.append(p)
    return unique


def skill_line(skills: Sequence[dict]) -> str:
    """'PyTorch or TensorFlow, Docker, SQL' — alternatives kept visible."""
    parts = []
    for s in skills or ():
        if isinstance(s, dict):
            name = s.get("name", "")
            alts = s.get("alternatives") or []
            parts.append(" or ".join([name] + list(alts)) if alts else name)
        else:
            parts.append(str(s))
    return ", ".join(p for p in parts if p)


def build_rows(path: Path) -> Tuple[str, List[dict], dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    job = data.get("job", {}) or {}
    role = job.get("role_title") or stem(str(path)).replace("_ranking", "")

    rows: List[dict] = []
    for c in list(data.get("shortlist", [])) + list(data.get("duplicates_folded", [])):
        comps = {k["label"]: k for k in c.get("components", [])}

        def bucket(label: str) -> str:
            k = comps.get(label)
            if not k:
                return ""
            return f"{k['earned']:.1f} / {k['possible']:.1f}\n{k.get('detail', '')}"

        assessment = "\n".join(
            f"{a['skill']} — {a['depth']}"
            + (f': "{a["evidence"][:90]}"' if a.get("evidence") else "")
            for a in c.get("assessments", []))

        exp = c.get("experience") or {}
        exp_text = ""
        if exp.get("years") is not None:
            exp_text = f"{exp['years']:g} yrs ({exp.get('method', '')})"
            if exp.get("evidence"):
                exp_text += "\n" + ", ".join(exp["evidence"][:4])

        rows.append({
            "resume": basename(c.get("path", "")),
            "jd": role,
            "score": c.get("score"),
            "rank": c.get("rank"),
            "matched": ", ".join(c.get("matched_skills") or []),
            "missing": ", ".join(c.get("missing_or_weak") or []),
            "mandatory": bucket("Mandatory skills"),
            "preferred": bucket("Preferred skills"),
            "assessment": assessment,
            "explanation": c.get("explanation", ""),
            "summary": c.get("summary", ""),
            "experience": exp_text,
            "flags": "\n".join(c.get("flags") or []),
            "conf": c.get("extraction_confidence"),
            "_duplicate": bool(c.get("duplicate_of")),
        })

    rows.sort(key=lambda r: (r["rank"] is None, r["rank"] or 0))
    return role, rows, job


# ---------------------------------------------------------------------------
# writing


def sheet_title(role: str, used: set) -> str:
    """Excel: 31 characters, and none of  [ ] : * ? / \\

    Forbidden characters are replaced rather than deleted -- dropping the slash
    turns "AI/ML Engineer" into "AIML Engineer", which reads like a typo.
    """
    swapped = "".join("-" if ch in "[]:*?/\\" else ch for ch in role)
    clean = " ".join(swapped.split())[:31].strip() or "Sheet"
    title, n = clean, 2
    while title.lower() in used:
        suffix = f" ({n})"
        title = clean[:31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def write_sheet(ws: Worksheet, rows: List[dict], job: dict) -> None:
    for col, (header, _, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28

    for r, row in enumerate(rows, start=2):
        for col, (_, key, _, wrap) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=col, value=row.get(key))
            cell.font = DUP_FONT if row["_duplicate"] else BODY_FONT
            cell.alignment = Alignment(
                vertical="top", wrap_text=wrap,
                horizontal="center" if key in ("score", "rank", "conf") else "left")
            cell.border = BORDER
            if key == "score":
                cell.number_format = "0.0"
            elif key == "conf":
                cell.number_format = "0.00"

    last = len(rows) + 1
    ws.freeze_panes = "C2"                    # keep resume + JD visible
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last}"
        # Green-to-red down the score column: the whole point of a sheet is
        # seeing the shape of a batch without reading every row.
        ws.conditional_formatting.add(
            f"C2:C{last}",
            ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="num", end_value=100, end_color="63BE7B"))

    note = ws.cell(row=last + 2, column=1,
                   value="Scores are computed deterministically from the depth "
                         "labels in Assessment; the model never emits a number. "
                         "Greyed italic rows are duplicate submissions folded "
                         "into a higher-ranked candidate.")
    note.font = Font(name=FONT, size=9, italic=True, color="808080")
    ws.cell(row=last + 3, column=1,
            value=f"Required: {skill_line(job.get('mandatory_skills'))}"
            ).font = Font(name=FONT, size=9, color="808080")
    ws.cell(row=last + 4, column=1,
            value=f"Preferred: {skill_line(job.get('preferred_skills'))}"
            ).font = Font(name=FONT, size=9, color="808080")


def write_summary(ws: Worksheet, sheets: List[Tuple[str, List[dict], dict]]) -> None:
    headers = ["Job Description", "Sheet", "Candidates", "Duplicates",
               "Top Score", "Median Score", "Scored 70+", "Scored 0",
               "Top Candidate"]
    widths = [30, 26, 12, 11, 10, 13, 11, 9, 30]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    for r, (title, rows, job) in enumerate(sheets, start=2):
        ranked = [x for x in rows if not x["_duplicate"]]
        scores = sorted((x["score"] for x in ranked if x["score"] is not None),
                        reverse=True)
        median = (scores[len(scores) // 2] if len(scores) % 2
                  else (scores[len(scores) // 2 - 1] + scores[len(scores) // 2]) / 2
                  ) if scores else None
        values = [job.get("role_title") or title, title, len(ranked),
                  len(rows) - len(ranked),
                  scores[0] if scores else None, median,
                  sum(1 for s in scores if s >= 70), sum(1 for s in scores if s == 0),
                  ranked[0]["resume"] if ranked else ""]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                horizontal="center" if 3 <= col <= 8 else "left")
            if col in (5, 6):
                cell.number_format = "0.0"
    ws.freeze_panes = "A2"


def export(paths: Sequence[Path], out: Path) -> Tuple[int, int]:
    built = [build_rows(p) for p in paths]

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    used: set = set()
    sheets = []
    for role, rows, job in built:
        title = sheet_title(role, used)
        write_sheet(wb.create_sheet(title), rows, job)
        sheets.append((title, rows, job))

    write_summary(summary, sheets)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return len(sheets), sum(len(r) for _, r, _ in built)


# ---------------------------------------------------------------------------


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        prog="export_excel.py",
        description="Export ranking results to a spreadsheet, one sheet per JD.")
    ap.add_argument("targets", nargs="+",
                    help="ranking JSON files, or a folder containing them")
    ap.add_argument("--out", default="rankings.xlsx")
    args = ap.parse_args(argv)

    paths = collect(args.targets)
    n_sheets, n_rows = export(paths, Path(args.out))

    print(f"wrote {args.out}")
    print(f"  {n_sheets} job description sheet(s), {n_rows} candidate row(s)")
    for p in paths:
        print(f"  from {p}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())